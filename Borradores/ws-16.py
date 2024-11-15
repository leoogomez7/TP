import requests as re
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


# Hacer la solicitud GET
response = re.get('https://www.digitalogy.co/blog/programming-languages-from-easy-to-hard/')
# Parsear el HTML con BeautifulSoup
soup = BeautifulSoup(response.text, 'html.parser')

# Buscar los elementos que contienen los lenguajes
language_section = soup.find_all('h4')  # Buscar todos los encabezados h2

# Crear una lista para almacenar los lenguajes de programación
languages = []

# Buscar los lenguajes dentro del contenido
for section in language_section:
    text = section.get_text(strip=True)
    # Identificar el texto relevante (en este caso los nombres de los lenguajes)
    if text and (text[0].isdigit() or text in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10','11', '12', '13', '14', '15', '16', '17', '18', '19', '20','21', '22', '23', '24', '25', '26', '27', '28', '29', '30']):
        languages.append(text)

# Verificar si se encontraron lenguajes
if languages:
    # Crear un DataFrame con los lenguajes
    df = pd.DataFrame(languages, columns=["Lenguajes de Programación de dificultad fácil a difícil"])

    # Nombre del archivo y título personalizado
    nombre_archivo = "Lenguajes de Programación de dificultad fácil a difícil.xlsx"
    titulo_personalizado = "Lenguajes de Programación de dificultad fácil a difícil"

    # Exportar a Excel
    df.to_excel(nombre_archivo, index=False, startrow=2)

    # Formatear el archivo con openpyxl
    wb = load_workbook(nombre_archivo)
    ws = wb.active

    # Insertar título en la primera fila y fusionar celdas
    ws["A1"] = titulo_personalizado
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)

    # Cambiar estilo del título
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Ajustar ancho de columna
    col_letter = get_column_letter(1)
    ws.column_dimensions[col_letter].width = max(15, len(titulo_personalizado) + 2)

    # Guardar los cambios
    wb.save(nombre_archivo)
    print(f"Datos exportados exitosamente a {nombre_archivo}")
else:
    print("No se encontraron lenguajes de programación.")