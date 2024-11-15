import requests as re
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# Hacer la solicitud GET
response1 = re.get('https://www.tiobe.com/tiobe-index/')

# Parsear el HTML con BeautifulSoup
soup1 = BeautifulSoup(response1.text, 'html.parser')

# Buscar la tabla en el HTML
table1 = soup1.find('table')

# Extraer las filas de la tabla
rows1 = table1.find_all('tr')

# Crear una lista para almacenar las filas
data1 = []

# Definir los títulos de las columnas personalizados
custom_headers1 = ["Rank Nov 2024", "Rank Nov 2023", "-", "-", "Lenguaje de Programación", "Calificación", "Cambio"]

# Agregar las filas de datos al archivo Excel
titulo_personalizado1 = "Lenguajes de programación más usados en 2024" # Título por defecto mostrado en pantalla "terminal"
print("\n" + titulo_personalizado1 + "\n" + "="*len(titulo_personalizado1) + "\n")
for row in rows1:
    cells = row.find_all(['td', 'th'])  # Obtener tanto las celdas de datos (td) como los encabezados (th)
    cell_text = [cell.text.strip() for cell in cells]  # Extraer texto de las celdas
    print(cell_text)
    data1.append(cell_text)  # Añadir cada fila a la lista de datos

# Exportar los datos a un archivo Excel
# Crear el DataFrame con las filas de datos (excluyendo el encabezado extraído, si existe)
df1 = pd.DataFrame(data1[1:], columns=custom_headers1)  # Usa data[1:] para omitir la primera fila si es el encabezado extraído

# Guardar el DataFrame en un archivo Excel
nombre_archivo1 = 'Lenguajes de programacion mas usados 2024.xlsx'
df1.to_excel(nombre_archivo1, index=False, startrow=2)  # Guardar, dejando espacio para el título

# Cargar el archivo con openpyxl para añadir el título en la primera fila
wb1 = load_workbook(nombre_archivo1)
ws1 = wb1.active

# Insertar el título en la primera celda y fusionar celdas para que abarque todas las columnas
ws1["A1"] = titulo_personalizado1
ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(custom_headers1))

# Cambiar la fuente y la alineación sin usar `copy`
ws1["A1"].font = Font(bold=True, size=14)
ws1["A1"].alignment = Alignment(horizontal="center")

# Ajustar el ancho de las columnas automáticamente
for col_num, col_title in enumerate(custom_headers1, 1):
    col_letter = get_column_letter(col_num)
    ws1.column_dimensions[col_letter].width = max(15, len(col_title) + 2)  # Ajusta según el contenido

# Guardar los cambios
wb1.save(nombre_archivo1)
