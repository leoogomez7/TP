import requests as re
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# Título por defecto mostrado en pantalla "terminal"
titulo_personalizado = "Lenguajes de programación más usados en 2024"
print("\n" + titulo_personalizado + "\n" + "="*len(titulo_personalizado) + "\n")

# Hacer la solicitud GET
response = re.get('https://www.tiobe.com/tiobe-index/')

# Parsear el HTML con BeautifulSoup
soup = BeautifulSoup(response.text, 'html.parser')

# Buscar la tabla en el HTML
table = soup.find('table')

# Extraer las filas de la tabla
rows = table.find_all('tr')

# Crear una lista para almacenar las filas
data = []

# Definir los títulos de las columnas personalizados
custom_headers = ["Rank Nov 2024", "Rank Nov 2023", "-", "-", "Lenguaje de Programación", "Calificación", "Cambio"]

# Agregar las filas de datos al archivo Excel
for row in rows:
    cells = row.find_all(['td', 'th'])  # Obtener tanto las celdas de datos (td) como los encabezados (th)
    cell_text = [cell.text.strip() for cell in cells]  # Extraer texto de las celdas
    print(cell_text)
    data.append(cell_text)  # Añadir cada fila a la lista de datos

# Exportar los datos a un archivo Excel
# Crear el DataFrame con las filas de datos (excluyendo el encabezado extraído, si existe)
df = pd.DataFrame(data[1:], columns=custom_headers)  # Usa data[1:] para omitir la primera fila si es el encabezado extraído

# Guardar el DataFrame en un archivo Excel
nombre_archivo = 'Lenguajes de programacion mas usados 2024.xlsx'
df.to_excel(nombre_archivo, index=False, startrow=2)  # Guardar, dejando espacio para el título

# Cargar el archivo con openpyxl para añadir el título en la primera fila
wb = load_workbook(nombre_archivo)
ws = wb.active

# Insertar el título en la primera celda y fusionar celdas para que abarque todas las columnas
ws["A1"] = titulo_personalizado
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(custom_headers))

# Cambiar la fuente y la alineación sin usar `copy`
ws["A1"].font = Font(bold=True, size=14)
ws["A1"].alignment = Alignment(horizontal="center")

# Ajustar el ancho de las columnas automáticamente
for col_num, col_title in enumerate(custom_headers, 1):
    col_letter = get_column_letter(col_num)
    ws.column_dimensions[col_letter].width = max(15, len(col_title) + 2)  # Ajusta según el contenido

# Guardar los cambios
wb.save(nombre_archivo)

#-----------------------------------------------------------
# Título para la segunda tabla
titulo_personalizado_2 = "Nueva tabla de datos"
print("\n" + titulo_personalizado_2 + "\n" + "="*len(titulo_personalizado_2) + "\n")

# Hacer la solicitud GET para la segunda página
response_2 = re.get('https://www.tiobe.com/tiobe-index/')  # NUEVO: Cambiar esta URL por la segunda página

# Parsear el HTML con BeautifulSoup para la segunda página
soup_2 = BeautifulSoup(response_2.text, 'html.parser')

# Buscar la tabla en el HTML de la segunda página
table_2 = soup_2.find('table')

# Extraer las filas de la tabla de la segunda página
rows_2 = table_2.find_all('tr')

# Crear una lista para almacenar las filas de la segunda tabla
data_2 = []

# Definir los títulos de las columnas personalizados para la segunda tabla
custom_headers_2 = ["Rank Nov 2024", "Rank Nov 2023", "-", "-", "Lenguaje de Programación", "Calificación", "Cambio"]  # NUEVO: Títulos para la segunda tabla

# Iterar sobre las filas de la segunda tabla
for row in rows_2:
    cells = row.find_all(['td', 'th'])  # Obtener tanto las celdas de datos (td) como los encabezados (th)
    cell_text = [cell.text.strip() for cell in cells]  # Extraer texto de las celdas
    data_2.append(cell_text)  # Añadir cada fila a la lista de datos de la segunda tabla

# Crear el DataFrame con las filas de la segunda tabla
df_2 = pd.DataFrame(data_2[1:], columns=custom_headers_2)

# Guardar los datos de la segunda tabla en un archivo Excel
nombre_archivo_2 = 'Nueva_tabla_de_datos.xlsx'
df_2.to_excel(nombre_archivo_2, index=False, startrow=2)

# Cargar el archivo de la segunda tabla con openpyxl para añadir el título
wb_2 = load_workbook(nombre_archivo_2)
ws_2 = wb_2.active

# Insertar el título y formato en la segunda tabla
ws_2["A1"] = titulo_personalizado_2
ws_2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(custom_headers_2))
ws_2["A1"].font = Font(bold=True, size=14)
ws_2["A1"].alignment = Alignment(horizontal="center")

# Ajustar el ancho de las columnas de la segunda tabla
for col_num, col_title in enumerate(custom_headers_2, 1):
    col_letter = get_column_letter(col_num)
    ws_2.column_dimensions[col_letter].width = max(15, len(col_title) + 2)

# Guardar los cambios en el archivo de la segunda tabla
wb_2.save(nombre_archivo_2)

#-----------------------------------------------------------
# Título para la segunda tabla
titulo_personalizado_3 = "Nueva tabla de datos"
print("\n" + titulo_personalizado_3 + "\n" + "="*len(titulo_personalizado_3) + "\n")

# Hacer la solicitud GET para la segunda página
response_3 = re.get('https://www.tiobe.com/tiobe-index/')  # NUEVO: Cambiar esta URL por la segunda página

# Parsear el HTML con BeautifulSoup para la segunda página
soup_3 = BeautifulSoup(response_3.text, 'html.parser')

# Buscar la tabla en el HTML de la segunda página
table_3 = soup_3.find('table')

# Extraer las filas de la tabla de la segunda página
rows_3 = table_3.find_all('tr')

# Crear una lista para almacenar las filas de la segunda tabla
data_3 = []

# Definir los títulos de las columnas personalizados para la segunda tabla
custom_headers_3 = ["Rank Nov 2024", "Rank Nov 2023", "-", "-", "Lenguaje de Programación", "Calificación", "Cambio"]  # NUEVO: Títulos para la segunda tabla

# Iterar sobre las filas de la segunda tabla
for row in rows_3:
    cells = row.find_all(['td', 'th'])  # Obtener tanto las celdas de datos (td) como los encabezados (th)
    cell_text = [cell.text.strip() for cell in cells]  # Extraer texto de las celdas
    data_3.append(cell_text)  # Añadir cada fila a la lista de datos de la segunda tabla

# Crear el DataFrame con las filas de la segunda tabla
df_3 = pd.DataFrame(data_3[1:], columns=custom_headers_3)

# Guardar los datos de la segunda tabla en un archivo Excel
nombre_archivo_3 = 'Nueva_tabla_de_datos33333.xlsx'
df_3.to_excel(nombre_archivo_2, index=False, startrow=2)

# Cargar el archivo de la segunda tabla con openpyxl para añadir el título
wb_3 = load_workbook(nombre_archivo_3)
ws_3 = wb_3.active

# Insertar el título y formato en la segunda tabla
ws_3["A1"] = titulo_personalizado_3
ws_3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(custom_headers_2))
ws_3["A1"].font = Font(bold=True, size=14)
ws_3["A1"].alignment = Alignment(horizontal="center")

# Ajustar el ancho de las columnas de la segunda tabla
for col_num, col_title in enumerate(custom_headers_3, 1):
    col_letter = get_column_letter(col_num)
    ws_3.column_dimensions[col_letter].width = max(15, len(col_title) + 2)

# Guardar los cambios en el archivo de la segunda tabla
wb_3.save(nombre_archivo_3)