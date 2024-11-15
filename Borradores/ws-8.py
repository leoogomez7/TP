import requests as re
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Título personalizado en pantalla
titulo_personalizado = "Lenguajes de programación más usados en 2024"
print("\n" + titulo_personalizado + "\n" + "="*len(titulo_personalizado) + "\n")

# Hacer la solicitud GET
response = re.get('https://www.tiobe.com/tiobe-index/')

# Parsear el HTML con BeautifulSoup
soup = BeautifulSoup(response.text, 'html.parser')

# Buscar la tabla en el HTML
table = soup.find('table')

# Extraer las filas de la tabla
rows = table.find_all('tr')

# Crear una lista para almacenar las filas
data = []

# Definir los títulos de las columnas personalizados
custom_headers = ["Ranking Noviembre 2024", "Ranking Noviembre 2023", "-", "-", "Lenguaje de Programación", "Calificación", "Cambio"]

# Agregar las filas de datos al archivo Excel
for row in rows:
    cells = row.find_all(['td', 'th'])
    cell_text = [cell.text.strip() for cell in cells]
    data.append(cell_text)

# Crear el DataFrame con las filas de datos (excluyendo el encabezado extraído, si existe)
df = pd.DataFrame(data[1:], columns=custom_headers)  # Usa data[1:] para omitir la primera fila si es el encabezado extraído

# Guardar el DataFrame en un archivo Excel
nombre_archivo = 'Lenguajes_de_programacion_mas_usados_2024.xlsx'
df.to_excel(nombre_archivo, index=False, startrow=2)  # Guardar, dejando espacio para el título

# Cargar el archivo con openpyxl para añadir el título en la primera fila
wb = load_workbook(nombre_archivo)
ws = wb.active

# Insertar el título en la primera celda, fusionar celdas para que abarque todas las columnas
ws["A1"] = titulo_personalizado
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(custom_headers))

# Dar formato al título
ws["A1"].font = ws["A1"].font.copy(bold=True, size=14)
ws["A1"].alignment = ws["A1"].alignment.copy(horizontal="center")

# Ajustar el ancho de las columnas automáticamente
for col_num, col_title in enumerate(custom_headers, 1):
    col_letter = get_column_letter(col_num)
    ws.column_dimensions[col_letter].width = max(15, len(col_title) + 2)  # Ajusta según el contenido

# Guardar los cambios
wb.save(nombre_archivo)
