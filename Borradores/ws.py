import requests as re
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# Hacer la solicitud GET de los 3 sitios web
response1 = re.get('https://www.tiobe.com/tiobe-index/')
response2 = re.get('https://golang.withcodeexample.com/blog/top-highest-paying-programming-languages-to-learn-in-2024/')
response3 = re.get('https://www.digitalogy.co/blog/programming-languages-from-easy-to-hard/')

# Parsear el HTML con BeautifulSoup de los 3 sitios web
soup1 = BeautifulSoup(response1.text, 'html.parser')
soup2 = BeautifulSoup(response2.text, 'html.parser')
soup3 = BeautifulSoup(response3.text, 'html.parser')

#--------------------------------------------------------------------------------------
#Sitio web 1 - Lenguajes de programación más usados en 2024
# Buscar la tabla en el HTML
table1 = soup1.find('table')

# Extraer las filas de la tabla
rows1 = table1.find_all('tr')

# Crear una lista para almacenar las filas
data1 = []

# Definir los títulos de las columnas personalizados
custom_headers1 = ["Rank Nov 2024", "Rank Nov 2023", "-", "-", "Lenguaje de Programación", "Calificación", "Cambio"]

# Agregar las filas de datos al archivo Excel
print("\n" + "=========================================")
titulo_personalizado1 = "RANKING Lenguajes de programación más usados en 2024:" # Título por defecto mostrado en pantalla "terminal"
print(titulo_personalizado1 + "\n")

for row in rows1:
    cells = row.find_all(['td', 'th'])  # Obtener tanto las celdas de datos (td) como los encabezados (th)
    cell_text = [cell.text.strip() for cell in cells]  # Extraer texto de las celdas
    print(cell_text)
    data1.append(cell_text)  # Añadir cada fila a la lista de datos

print("\n" + "=========================================")
# Exportar los datos a un archivo Excel
# Crear el DataFrame con las filas de datos (excluyendo el encabezado extraído, si existe)
df1 = pd.DataFrame(data1[1:], columns=custom_headers1)  # Usa data[1:] para omitir la primera fila si es el encabezado extraído

# Guardar el DataFrame en un archivo Excel
nombre_archivo1 = 'Lenguajes de programacion mas usados 2024.xlsx'
df1.to_excel(nombre_archivo1, index=False, startrow=2)  # Guardar, dejando espacio para el título

# Cargar el archivo con openpyxl para añadir el título en la primera fila
wb1 = load_workbook(nombre_archivo1)
ws1 = wb1.active

# Insertar el título en la primera celda y fusionar celdas para que abarque todas las columnas
ws1["A1"] = titulo_personalizado1
ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(custom_headers1))

# Cambiar la fuente y la alineación sin usar `copy`
ws1["A1"].font = Font(bold=True, size=14)
ws1["A1"].alignment = Alignment(horizontal="center")

# Ajustar el ancho de las columnas automáticamente
for col_num, col_title in enumerate(custom_headers1, 1):
    col_letter = get_column_letter(col_num)
    ws1.column_dimensions[col_letter].width = max(15, len(col_title) + 2)  # Ajusta según el contenido

# Guardar los cambios
    wb1.save(nombre_archivo1)
#--------------------------------------------------------------------------------------
#Sitio web 2 - Lenguajes de Programación que mejor pagan en 2024
# Buscar los elementos que contienen los lenguajes
language_section2 = soup2.find_all('h2')  # Buscar todos los encabezados h2

# Crear una lista para almacenar los lenguajes de programación
languages2 = []

# Buscar los lenguajes dentro del contenido
for section in language_section2:
    text = section.get_text(strip=True)
    # Identificar el texto relevante (en este caso los nombres de los lenguajes)
    if text and (text[0].isdigit() or text in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10']):
        languages2.append(text)
    
    # Verificar si se encontraron lenguajes
if languages2:
    languages2.reverse()
 
    # Crear un DataFrame con los lenguajes
    df2 = pd.DataFrame(languages2, columns=["RANKING Lenguajes de Programación que mejor pagan en 2024:"])

    # Mostrar el DataFrame en la terminal
    #print("\n")
    print(df2.to_string(index=False, justify='left'))
    print("\n"+"========================================="+"\n")

    # Nombre del archivo y título personalizado
    nombre_archivo2 = "Lenguajes de Programación que mejor pagan en 2024.xlsx"

    # Exportar a Excel
    df2.to_excel(nombre_archivo2, index=False)#, startrow=2)

    # Formatear el archivo con openpyxl
    wb2 = load_workbook(nombre_archivo2)
    ws2 = wb2.active

    # Insertar título en la primera fila y fusionar celdas
    num_columns = len(df2.columns)  # Contar el número de columnas en el DataFrame

    # Ajustar ancho de columna
    for col_num, col_title in enumerate(df2.columns, 1):
        col_letter = get_column_letter(col_num)
        ws2.column_dimensions[col_letter].width = max(15, len(col_title) + 2)

    # Guardar los cambios
    wb2.save(nombre_archivo2)
#--------------------------------------------------------------------------------------
#Sitio web 3 - Lenguajes de Programación ordenados de fácil a difícil según su complejidad.
# Buscar los elementos que contienen los lenguajes
language_section3 = soup3.find_all('h4')  # Buscar todos los encabezados h2

# Crear una lista para almacenar los lenguajes de programación
languages3 = []

# Buscar los lenguajes dentro del contenido
for section in language_section3:
    text = section.get_text(strip=True)
    # Identificar el texto relevante (en este caso los nombres de los lenguajes)
    if text and (text[0].isdigit() or text in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10','11', '12', '13', '14', '15', '16', '17', '18', '19', '20','21', '22', '23', '24', '25', '26', '27', '28', '29', '30']):
        languages3.append(text)
    
    # Verificar si se encontraron lenguajes
if languages3:

    # Crear un DataFrame con los lenguajes
    df3 = pd.DataFrame(languages3, columns=["RANKING Lenguajes de Programación ordenados de fácil a difícil según su complejidad:"])

    # Mostrar el DataFrame en la terminal
    #print("\n")
    print(df3.to_string(index=False, justify='left'))
    print("\n"+"========================================="+"\n")

    # Nombre del archivo y título personalizado
    nombre_archivo3 = "Lenguajes de Programación ordenados de fácil a difícil según su complejidad.xlsx"

    # Exportar a Excel
    df3.to_excel(nombre_archivo3, index=False)#, startrow=2)

    # Formatear el archivo con openpyxl
    wb3 = load_workbook(nombre_archivo3)
    ws3 = wb3.active

    # Insertar título en la primera fila y fusionar celdas
    num_columns = len(df3.columns)  # Contar el número de columnas en el DataFrame

    # Ajustar ancho de columna
    for col_num, col_title in enumerate(df2.columns, 1):
        col_letter = get_column_letter(col_num)
        ws3.column_dimensions[col_letter].width = max(15, len(col_title) + 2)

    # Guardar los cambios
    wb3.save(nombre_archivo3)