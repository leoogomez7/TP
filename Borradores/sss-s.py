import requests as re
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Hacer la solicitud GET de los 3 sitios web
response1 = re.get('https://www.tiobe.com/tiobe-index/')
response2 = re.get('https://golang.withcodeexample.com/blog/top-highest-paying-programming-languages-to-learn-in-2024/')
response3 = re.get('https://www.digitalogy.co/blog/programming-languages-from-easy-to-hard/')

# Parsear el HTML con BeautifulSoup de los 3 sitios web
soup1 = BeautifulSoup(response1.text, 'html.parser')
soup2 = BeautifulSoup(response2.text, 'html.parser')
soup3 = BeautifulSoup(response3.text, 'html.parser')

# Crear un archivo de Excel nuevo
wb = Workbook()

#--------------------------------------------------------------------------------------
# Sitio web 1 - Lenguajes de programación más usados en 2024
# Buscar la tabla en el HTML
table1 = soup1.find('table')
rows1 = table1.find_all('tr')
data1 = []

custom_headers1 = ["Rank Nov 2024", "Rank Nov 2023", "-", "-", "Lenguaje de Programación", "Calificación", "Cambio"]

for row in rows1:
    cells = row.find_all(['td', 'th'])  # Obtener tanto las celdas de datos (td) como los encabezados (th)
    cell_text = [cell.text.strip() for cell in cells]  # Extraer texto de las celdas
    data1.append(cell_text)

# Crear una nueva hoja en el archivo Excel
ws1 = wb.create_sheet(title="Lenguajes más usados 2024")

# Escribir el título y los datos
ws1["A1"] = "RANKING Lenguajes de programación más usados en 2024:"
ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(custom_headers1))
ws1["A1"].font = Font(bold=True, size=14)
ws1["A1"].alignment = Alignment(horizontal="center")

# Escribir los encabezados
for col_num, col_title in enumerate(custom_headers1, 1):
    ws1.cell(row=2, column=col_num, value=col_title)

# Escribir los datos
for row_num, row_data in enumerate(data1[1:], 3):  # Excluir el encabezado extraído
    for col_num, cell_value in enumerate(row_data, 1):
        ws1.cell(row=row_num, column=col_num, value=cell_value)

# Ajustar el ancho de las columnas
for col_num, col_title in enumerate(custom_headers1, 1):
    col_letter = get_column_letter(col_num)
    ws1.column_dimensions[col_letter].width = max(15, len(col_title) + 2)

#--------------------------------------------------------------------------------------
# Sitio web 2 - Lenguajes de Programación que mejor pagan en 2024
language_section2 = soup2.find_all('h2')
languages2 = [section.get_text(strip=True) for section in language_section2 if section.get_text(strip=True).isdigit()]

# Crear una nueva hoja en el archivo Excel
ws2 = wb.create_sheet(title="Lenguajes mejor pagos 2024")

# Escribir el título y los datos
ws2["A1"] = "RANKING Lenguajes de Programación que mejor pagan en 2024:"
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)
ws2["A1"].font = Font(bold=True, size=14)
ws2["A1"].alignment = Alignment(horizontal="center")

# Escribir los lenguajes
for row_num, language in enumerate(languages2, 2):
    ws2.cell(row=row_num, column=1, value=language)

#--------------------------------------------------------------------------------------
# Sitio web 3 - Lenguajes de Programación de dificultad fácil a difícil
language_section3 = soup3.find_all('h4')
languages3 = [section.get_text(strip=True) for section in language_section3 if section.get_text(strip=True).isdigit()]

# Crear una nueva hoja en el archivo Excel
ws3 = wb.create_sheet(title="Lenguajes por dificultad 2024")

# Escribir el título y los datos
ws3["A1"] = "RANKING Lenguajes de Programación de dificultad fácil a difícil:"
ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)
ws3["A1"].font = Font(bold=True, size=14)
ws3["A1"].alignment = Alignment(horizontal="center")

# Escribir los lenguajes
for row_num, language in enumerate(languages3, 2):
    ws3.cell(row=row_num, column=1, value=language)

# Guardar el archivo de Excel
wb.save("Lenguajes_de_Programacion_2024.xlsx")
