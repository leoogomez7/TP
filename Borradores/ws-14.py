import requests as re
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


# Hacer la solicitud GET
response = re.get('https://golang.withcodeexample.com/blog/top-highest-paying-programming-languages-to-learn-in-2024/')
# Parsear el HTML con BeautifulSoup
soup = BeautifulSoup(response.text, 'html.parser')

# Buscar los elementos que contienen los lenguajes
language_section = soup.find_all('h2')  # Buscar todos los encabezados h2

# Crear una lista para almacenar los lenguajes de programación
languages = []

# Buscar los lenguajes dentro del contenido
for section in language_section:
    text = section.get_text(strip=True)
    # Identificar el texto relevante (en este caso los nombres de los lenguajes)
    if text and (text[0].isdigit() or text in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10']):
        languages.append(text)

# Verificar si se encontraron lenguajes
if languages:
    languages.reverse()
    # Crear un DataFrame con los lenguajes
    df = pd.DataFrame(languages, columns=["Lenguajes de Programación Más Altamente Pagados en 2024"])

    # Nombre del archivo y título personalizado
    nombre_archivo = "Lenguajes_Mas_Altamente_Pagados_2024.xlsx"
    titulo_personalizado = "Top Lenguajes de Programación Más Altamente Pagados en 2024"

    # Exportar a Excel
    df.to_excel(nombre_archivo, index=False, startrow=2)

    # Formatear el archivo con openpyxl
    wb = load_workbook(nombre_archivo)
    ws = wb.active

    # Insertar título en la primera fila y fusionar celdas
    ws["A1"] = titulo_personalizado
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)

    # Cambiar estilo del título
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Ajustar ancho de columna
    col_letter = get_column_letter(1)
    ws.column_dimensions[col_letter].width = max(15, len(titulo_personalizado) + 2)

    # Guardar los cambios
    wb.save(nombre_archivo)
    print(f"Datos exportados exitosamente a {nombre_archivo}")
else:
    print("No se encontraron lenguajes de programación.")