import requests as re
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# Hace la solicitud GET de los 3 sitios web
response1 = re.get('https://www.tiobe.com/tiobe-index/')
response2 = re.get('https://golang.withcodeexample.com/blog/top-highest-paying-programming-languages-to-learn-in-2024/')
response3 = re.get('https://www.digitalogy.co/blog/programming-languages-from-easy-to-hard/')

# Parsea el HTML con BeautifulSoup de los 3 sitios web
soup1 = BeautifulSoup(response1.text, 'html.parser')
soup2 = BeautifulSoup(response2.text, 'html.parser')
soup3 = BeautifulSoup(response3.text, 'html.parser')

#--------------------------------------------------------------------------------------
#Sitio web 1 - Lenguajes de programación más usados en 2024
# Busca la tabla en el HTML
table1 = soup1.find('table')

# Extrae las filas de la tabla
rows1 = table1.find_all('tr')

# Crea una lista para almacenar las filas
data1 = []

# Define los títulos de las columnas personalizados
custom_headers = ["Rank Nov 2024", "Rank Nov 2023", "-", "-", "Lenguaje de Programación", "Calificación", "Cambio"]

#Título por defecto mostrado en pantalla "terminal"
print("\n" + "=========================================")
titulo_personalizado1 = "RANKING Lenguajes de programación más usados en 2024:"
print(titulo_personalizado1 + "\n")

# Agrega las filas de datos al archivo Excel
for row in rows1:
    cells = row.find_all(['td', 'th'])  # Obtiene tanto las celdas de datos (td) como los encabezados (th)
    cell_text = [cell.text.strip() for cell in cells]  # Extrae texto de las celdas
    print(cell_text)
    data1.append(cell_text)  # Añade cada fila a la lista de datos

print("\n" + "=========================================")

# Exporta los datos a un archivo Excel
# Crea el DataFrame con las filas de datos (excluyendo el encabezado extraído, si existe)
df1 = pd.DataFrame(data1[1:], columns=custom_headers)  # Usa data[1:] para que omita la primera fila si es el encabezado extraído
# Guarda el DataFrame en un archivo Excel
nombre_archivo1 = 'LP - mas usados 2024.xlsx'
df1.to_excel(nombre_archivo1, index=False, startrow=2)
# Carga el archivo con openpyxl para añadir el título en la primera fila
wb1 = load_workbook(nombre_archivo1)
ws1 = wb1.active

# Inserta el título en la primera celda y fusiona celdas para que abarque todas las columnas
ws1["A1"] = titulo_personalizado1
ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(custom_headers))

# Cambia la fuente y la alineación
ws1["A1"].font = Font(bold=True, size=14)
ws1["A1"].alignment = Alignment(horizontal="center")

# Ajusta el ancho de las columnas automáticamente
for col_num, col_title in enumerate(custom_headers, 1):
    col_letter = get_column_letter(col_num)
    ws1.column_dimensions[col_letter].width = max(15, len(col_title) + 2)

# Guarda los cambios
    wb1.save(nombre_archivo1)
#--------------------------------------------------------------------------------------
#Sitio web 2 - Lenguajes de Programación que mejor pagan en 2024
# Busca los elementos que contienen los lenguajes
language_section2 = soup2.find_all('h2')  # Busca todos los encabezados h2

# Crea una lista para almacenar los lenguajes de programación
languages2 = []

# Busca los lenguajes de programación dentro del contenido
for section in language_section2:
    text = section.get_text(strip=True)
    # Identifica los nombres de los lenguajes
    if text and (text[0].isdigit() or text in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10']):
        languages2.append(text)
    
    # Verifica si se encontraron lenguajes
if languages2:
    
    #Invierte los lenguajes porque en el sitio web estan ordenado alreves
    languages2.reverse()
    
    # Crea un DataFrame con los lenguajes
    df2 = pd.DataFrame(languages2, columns=["RANKING Lenguajes de Programación que mejor pagan en 2024:"])
    
    # Muestra el DataFrame en la terminal
    print(df2.to_string(index=False, justify='left'))
    print("\n"+"========================================="+"\n")

    # Nombre del archivo y título personalizado
    nombre_archivo2 = "LP - mejor pagan en 2024.xlsx"

    # Exporta a Excel
    df2.to_excel(nombre_archivo2, index=False)

    # Formatea el archivo con openpyxl
    wb2 = load_workbook(nombre_archivo2)
    ws2 = wb2.active

    # Inserta título en la primera fila y fusiona celdas
    num_columns = len(df2.columns)  # Cuenta el número de columnas en el DataFrame

    # Ajusta ancho de columna
    for col_num, col_title in enumerate(df2.columns, 1):
        col_letter = get_column_letter(col_num)
        ws2.column_dimensions[col_letter].width = max(15, len(col_title) + 2)

    # Guarda los cambios
    wb2.save(nombre_archivo2)
#--------------------------------------------------------------------------------------
#Sitio web 3 - Lenguajes de Programación ordenados de fácil a difícil según su complejidad.
# Busca los elementos que contienen los lenguajes
language_section3 = soup3.find_all('h4')  # Busca todos los encabezados h2

# Crea una lista para almacenar los lenguajes de programación
languages3 = []

#Busca los lenguajes dentro del contenido
for section in language_section3:
    text = section.get_text(strip=True)
    #Identifica los nombres de los lenguajes de programación
    if text and (text[0].isdigit() or text in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10','11', '12', '13', '14', '15', '16', '17', '18', '19', '20','21', '22', '23', '24', '25', '26', '27', '28', '29', '30']):
        languages3.append(text)
    
#Verifica si se encontraron los lenguajes de programación
if languages3:

    #Crea un DataFrame con los lenguajes de programación
    df3 = pd.DataFrame(languages3, columns=["RANKING Lenguajes de Programación ordenados de fácil a difícil según su complejidad:"])

    #Muestra el DataFrame en la terminal
    print(df3.to_string(index=False, justify='left'))
    print("\n"+"========================================="+"\n")

    # Nombre del archivo y título personalizado
    nombre_archivo3 = "LP - complejidad.xlsx"

    # Exporta a Excel
    df3.to_excel(nombre_archivo3, index=False)

    #Formatea el archivo con openpyxl
    wb3 = load_workbook(nombre_archivo3)
    ws3 = wb3.active

    # Inserta título en la primera fila y fusiona celdas
    num_columns = len(df3.columns)  # Cuenta el número de columnas en el DataFrame

    # Ajusta ancho de columna
    for col_num, col_title in enumerate(df2.columns, 1):
        col_letter = get_column_letter(col_num)
        ws3.column_dimensions[col_letter].width = max(15, len(col_title) + 2)

    # Guarda los cambios
    wb3.save(nombre_archivo3)
#--------------------------------------------------------------------------------------
# Guarda en un archivo de Excel con tres libros los 3 ranking's
with pd.ExcelWriter('Los 3 rankings de LP.xlsx') as writer:
    df1.to_excel(writer, sheet_name='Ranking - más usados 2024', index=False)
    df2.to_excel(writer, sheet_name='Ranking - mejor pagan en 2024', index=False)
    df3.to_excel(writer, sheet_name='Ranking - complejidad', index=False)

#Muestra por pantalla "terminal" de que se generaron los archivos
print("Archivos Excel generados"+ "\n")
